import frappe
from frappe.utils import get_files_path, get_site_path, get_site_base_path
import openpyxl,re


def get_absolute_path(file_name, is_private=False):
	site_name = get_site_base_path()
	if(file_name.startswith('/files/')):
		file_name = file_name[7:]
	return frappe.utils.get_bench_path()+ "/sites/"+site_name[2:]+"/"+ frappe.utils.get_path('private' if is_private else 'public', 'files', file_name)[1:]

@frappe.whitelist()
def get_column_names(file_name):
	excel_file_path = get_absolute_path(file_name)
	wb = openpyxl.load_workbook(excel_file_path)
	sheet = wb.active
	max_col = sheet.max_column
	column_map = []
	
	for i in range(1,max_col+1):
		cell_obj = sheet.cell(row=1,column=i)
		column_map.append({
			"excel_column" : cell_obj.value,
			"import_column": ""
		})
	return column_map